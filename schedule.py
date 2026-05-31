import os
import sys
import shutil
from datetime import datetime, date, timedelta
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, "frozen", False) else __file__))
EXCEL_FILE = os.path.join(BASE_DIR, "programma_ypiresion_step1.xlsm")
LOG_FILE = os.path.join(BASE_DIR, "scheduler_log.txt")


def log(msg):
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(msg + "\n")
RESERVE_GAP_DAYS = 3

REGULAR_SERVICE_COLS = list(range(3, 12))   # ΜΛΣΝ1..ΒΑΥΔΜ
RESERVE_SERVICE_COLS = [12, 13]             # ΕΠΙΦ_ΜΛΣΝ, ΕΠΙΦ_ΒΑΥΔΜ

DAY_NAMES_GR = {
    0: "Δευτέρα",
    1: "Τρίτη",
    2: "Τετάρτη",
    3: "Πέμπτη",
    4: "Παρασκευή",
    5: "Σάββατο",
    6: "Κυριακή",
}

# Κανονικοποίηση ημερών: αποδέχεται κεφαλαία, χωρίς τόνο, συντομογραφίες κτλ.
_DAY_NORMALIZE = {
    "ΔΕΥΤΕΡΑ": "Δευτέρα", "ΔΕΥΤ": "Δευτέρα",
    "ΤΡΙΤΗ": "Τρίτη", "ΤΡΙΤ": "Τρίτη",
    "ΤΕΤΑΡΤΗ": "Τετάρτη", "ΤΕΤ": "Τετάρτη",
    "ΠΕΜΠΤΗ": "Πέμπτη", "ΠΕΜΠ": "Πέμπτη",
    "ΠΑΡΑΣΚΕΥΗ": "Παρασκευή", "ΠΑΡΑΣΚ": "Παρασκευή", "ΠΑΡ": "Παρασκευή",
    "ΣΑΒΒΑΤΟ": "Σάββατο", "ΣΑΒ": "Σάββατο",
    "ΚΥΡΙΑΚΗ": "Κυριακή", "ΚΥΡ": "Κυριακή",
}

def normalize_day(day_str):
    """Μετατρέπει οποιοδήποτε format ημέρας στο κανονικό (π.χ. ΤΡΙΤΗ -> Τρίτη)."""
    import unicodedata
    # Αφαίρεση τόνων + κεφαλαία
    s = unicodedata.normalize("NFD", day_str.strip())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn").upper()
    return _DAY_NORMALIZE.get(s, day_str.strip())

SERVICE_COLUMNS = {
    "ΜΛΣΝ1": 3,
    "ΜΛΣΝ2": 4,
    "ΜΛΣΝ3": 5,
    "ΜΛΣΝ4": 6,
    "ΕΤΑΦ1": 7,
    "ΕΤΑΦ2": 8,
    "ΕΣΤ1": 9,
    "ΕΣΤ2": 10,
    "ΒΑΥΔΜ": 11,
    "ΕΠΙΦ_ΜΛΣΝ": 12,
    "ΕΠΙΦ_ΒΑΥΔΜ": 13,
}

REGULAR_SERVICE_SLOTS = {
    "mlsn": ["ΜΛΣΝ1", "ΜΛΣΝ2", "ΜΛΣΝ3", "ΜΛΣΝ4"],
    "etaf": ["ΕΤΑΦ1", "ΕΤΑΦ2"],
    "vavdm": ["ΒΑΥΔΜ"],
    "est": ["ΕΣΤ1", "ΕΣΤ2"],
}


def normalize_int(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def is_weekend(day_name):
    return day_name in ["Σάββατο", "Κυριακή"]


def get_week_number(current_date):
    return current_date.isocalendar()[1]


def read_personnel(ws_personnel):
    people = []

    for row in ws_personnel.iter_rows(min_row=2, values_only=True):
        name = row[0]

        if name is None or str(name).strip() == "":
            continue

        person = {
            "name": str(name).strip(),
            "posto": "" if row[1] is None else str(row[1]).strip(),
            "mlsn": normalize_int(row[2]),
            "etaf": normalize_int(row[3]),
            "est": normalize_int(row[4]),
            "vavdm": normalize_int(row[5]),
            "epif_mlsn": normalize_int(row[6]),
            "epif_vavdm": normalize_int(row[7]),
            "kanonikos": normalize_int(row[8]),
            "mono_1_ton_mina": normalize_int(row[9]),
            "mono_sk": normalize_int(row[10]),
            "energos": normalize_int(row[11], default=1),
            "max_mina": normalize_int(row[12]),
            "max_sk": normalize_int(row[13]),
            "min_gap_days": normalize_int(row[14], default=5),
            "sxolia": "" if row[15] is None else str(row[15]).strip(),
            "allowed_days": [] if (len(row) <= 16 or row[16] is None or str(row[16]).strip() == "")
                            else [normalize_day(d) for d in str(row[16]).split(",") if d.strip()],
            # Per-service max (0 = χωρίς όριο)
            "max_mlsn":  normalize_int(row[17]) if len(row) > 17 else 0,
            "max_etaf":  normalize_int(row[18]) if len(row) > 18 else 0,
            "max_est":   normalize_int(row[19]) if len(row) > 19 else 0,
            "max_vavdm": normalize_int(row[20]) if len(row) > 20 else 0,

            # κανονικές υπηρεσίες
            "last_service_date": None,
            "last_service_week": None,
            "regular_count": 0,
            "weekend_count": 0,
            "mlsn_count": 0,
            "etaf_count": 0,
            "est_count": 0,
            "vavdm_count": 0,

            # επιφυλακές
            "last_reserve_date": None,
            "reserve_count": 0,
        }

        people.append(person)

    return people


def read_needs(ws_needs):
    needs_by_day = {}

    for row in ws_needs.iter_rows(min_row=2, values_only=True):
        day_name = row[0]

        if day_name is None or str(day_name).strip() == "":
            continue

        day_key = str(day_name).strip()

        needs_by_day[day_key] = {
            "mlsn": normalize_int(row[1]),
            "etaf": normalize_int(row[2]),
            "est": normalize_int(row[3]),
            "vavdm": normalize_int(row[4]),
            "epif_mlsn": normalize_int(row[5]),
            "epif_vavdm": normalize_int(row[6]),
        }

    return needs_by_day


def to_python_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y",
                    "%d-%b-%Y", "%d-%b", "%d/%b/%Y", "%d/%b"):
            try:
                d = datetime.strptime(value, fmt)
                # Αν δεν έχει χρόνο (π.χ. "7-Apr"), βάζουμε τρέχον έτος
                if d.year == 1900:
                    d = d.replace(year=datetime.today().year)
                return d.date()
            except ValueError:
                continue
        raise ValueError(f"Δεν αναγνωρίζεται η ημερομηνία: '{value}'")
    raise ValueError(f"Μη έγκυρη ημερομηνία: {value} (τύπος: {type(value)})")


def read_schedule_rows(ws_schedule):
    expected_headers = [
        "ΗΜΕΡΑ",
        "ΗΜ/ΝΙΑ",
        "ΜΛΣΝ1",
        "ΜΛΣΝ2",
        "ΜΛΣΝ3",
        "ΜΛΣΝ4",
        "ΕΤΑΦ1",
        "ΕΤΑΦ2",
        "ΕΣΤ1",
        "ΕΣΤ2",
        "ΒΑΥΔΜ",
        "ΕΠΙΦ_ΜΛΣΝ",
        "ΕΠΙΦ_ΒΑΥΔΜ",
    ]

    actual_headers = []
    for col in range(1, len(expected_headers) + 1):
        actual_headers.append(ws_schedule.cell(row=3, column=col).value)

    if actual_headers != expected_headers:
        raise ValueError(
            f"Οι στήλες του Schedule δεν είναι σωστές.\n"
            f"Expected: {expected_headers}\n"
            f"Actual:   {actual_headers}"
        )

    month_start_raw = ws_schedule["B1"].value
    month_start = to_python_date(month_start_raw)

    year = month_start.year
    month = month_start.month
    days_in_month = calendar.monthrange(year, month)[1]

    schedule_rows = []

    for day_num in range(1, days_in_month + 1):
        current_date = date(year, month, day_num)
        weekday_index = current_date.weekday()
        day_name_gr = DAY_NAMES_GR[weekday_index]

        row_index = 3 + day_num

        schedule_rows.append({
            "row_index": row_index,
            "day_num": day_num,
            "day_name": day_name_gr,
            "date": current_date,
        })

    return schedule_rows


def clear_schedule_services(ws_schedule, schedule_rows):
    for row_info in schedule_rows:
        row_idx = row_info["row_index"]

        for col_idx in SERVICE_COLUMNS.values():
            ws_schedule.cell(row=row_idx, column=col_idx).value = None


def posto_allows_service(person, service, day_name):
    posto = person["posto"]

    if posto == "" or posto == "ΓΕΝΙΚΟ":
        return True

    weekend = is_weekend(day_name)

    # ΚΗΠΟΥΡΟΣ:
    # καθημερινές -> όχι ΜΛΣΝ, όχι ΕΣΤ
    # καθημερινές -> ναι ΕΤΑΦ, ναι ΒΑΥΔΜ
    # ΣΚ -> επιτρέπονται όλα
    if posto == "ΚΗΠΟΥΡΟΣ":
        if not weekend and service in ["mlsn", "est"]:
            return False

    if posto == "ΓΔΥ (ΚΩΣΤΟΥΛΑ)":
        if not weekend and service in ["mlsn", "est"]:
            return False

    return True


def respects_min_gap(person, current_date):
    last_date = person.get("last_service_date")

    if last_date is None:
        return True

    days_diff = (current_date - last_date).days
    return days_diff >= person["min_gap_days"]


def respects_weekly_limit(person, current_date):
    # Μόνο για κανονικούς
    if person["kanonikos"] != 1:
        return True

    current_week = get_week_number(current_date)
    last_week = person.get("last_service_week")

    if last_week is None:
        return True

    return current_week != last_week


def respects_max_weekend(person, day_name):
    if not is_weekend(day_name):
        return True

    return person["weekend_count"] < person["max_sk"]


def respects_reserve_gap(person, current_date):
    """
    Πιο ήπιο gap για επιφυλακές.
    Δεν θέλουμε κάποιος να έχει κανονική υπηρεσία και αμέσως μετά
    επιφυλακή ή το ανάποδο με πολύ μικρή απόσταση.
    """
    last_regular = person.get("last_service_date")
    last_reserve = person.get("last_reserve_date")

    if last_regular is not None:
        if (current_date - last_regular).days < RESERVE_GAP_DAYS:
            return False

    if last_reserve is not None:
        if (current_date - last_reserve).days < RESERVE_GAP_DAYS:
            return False

    return True


def is_special_weekend_only_once(person):
    return person["mono_1_ton_mina"] == 1 and person["mono_sk"] == 1


def service_flex(person):
    return (
        person["mlsn"]
        + person["etaf"]
        + person["est"]
        + person["vavdm"]
        + person["epif_mlsn"]
        + person["epif_vavdm"]
    )


def candidate_priority(service, person, day_name):
    weekend = is_weekend(day_name)

    # ΜΟΝΟ_1_ΤΟΝ_ΜΗΝΑ + ΜΟΝΟ_ΣΚ να προτιμώνται σε ΣΚ
    preferred_one_month_weekend = 0 if (weekend and is_special_weekend_only_once(person)) else 1

    # Για ΒΑΥΔΜ/ΕΣΤ θέλουμε πιο περιορισμένους
    if service in ["vavdm", "est"]:
        return (
            preferred_one_month_weekend,
            service_flex(person),
            person["regular_count"],
            person["weekend_count"],
            person["name"],
        )

    # Για MLSN/ETAF fairness
    return (
        preferred_one_month_weekend,
        person["regular_count"],
        person["weekend_count"],
        person["name"],
    )


def reserve_priority(service, person):
    # Για ΕΠΙΦ_ΒΑΥΔΜ προτιμάμε πιο περιορισμένους
    if service == "epif_vavdm":
        return (
            service_flex(person),
            person["reserve_count"],
            person["name"],
        )

    return (
        person["reserve_count"],
        person["name"],
    )


def service_allowed_flag(person, service):
    return person[service] == 1


def respects_service_max(person, service):
    """Ελέγχει per-service max (π.χ. max_etaf=1). 0 = χωρίς όριο."""
    max_key = f"max_{service}"
    limit = person.get(max_key, 0)
    if limit <= 0:
        return True
    count_key = f"{service}_count"
    return person[count_key] < limit


def find_regular_candidate(people, service, day_name, current_date, assigned_today, start_index, scheduled_dates=None, reserve_scheduled_dates=None, debug_log=False):
    eligible = []
    rejection_reasons = {} if debug_log else None

    for p in people:
        name = p["name"]
        if p["energos"] != 1:
            if debug_log: rejection_reasons[name] = "energos=0"
            continue

        if not service_allowed_flag(p, service):
            if debug_log: rejection_reasons[name] = f"{service}=0 (flag)"
            continue

        if not respects_service_max(p, service):
            if debug_log: rejection_reasons[name] = f"max_{service} φτασε"
            continue

        if name in assigned_today:
            if debug_log: rejection_reasons[name] = "ήδη σήμερα"
            continue

        if p["mono_sk"] == 1 and not is_weekend(day_name):
            if debug_log: rejection_reasons[name] = "mono_sk καθημερινή"
            continue

        if p["mono_1_ton_mina"] == 1 and p["regular_count"] >= 1:
            if debug_log: rejection_reasons[name] = "mono_1_ton_mina εξαντλήθηκε"
            continue

        if p["regular_count"] >= p["max_mina"]:
            if debug_log: rejection_reasons[name] = f"max_mina={p['max_mina']} φτασε ({p['regular_count']})"
            continue

        if p["allowed_days"] and day_name not in p["allowed_days"]:
            if debug_log: rejection_reasons[name] = f"allowed_days {p['allowed_days']}"
            continue

        if not posto_allows_service(p, service, day_name):
            if debug_log: rejection_reasons[name] = f"posto '{p['posto']}' απαγορεύει"
            continue

        if not respects_min_gap(p, current_date):
            if debug_log: rejection_reasons[name] = f"min_gap: τελευταία {p.get('last_service_date')} (gap<{p['min_gap_days']})"
            continue

        if not respects_reserve_gap(p, current_date):
            if debug_log: rejection_reasons[name] = f"reserve_gap: τελευταία επιφ {p.get('last_reserve_date')} (gap<{RESERVE_GAP_DAYS})"
            continue

        if not respects_weekly_limit(p, current_date):
            if debug_log: rejection_reasons[name] = "weekly limit"
            continue

        if not respects_max_weekend(p, day_name):
            if debug_log: rejection_reasons[name] = f"max_sk={p['max_sk']} φτασε ({p['weekend_count']})"
            continue

        if scheduled_dates is not None:
            person_dates = scheduled_dates.get(name, set())
            close_dates = [d for d in person_dates if abs((current_date - d).days) < p["min_gap_days"]]
            if close_dates:
                if debug_log: rejection_reasons[name] = f"forward gap: υπηρεσία σε {sorted(close_dates)} (min_gap={p['min_gap_days']})"
                continue

        if reserve_scheduled_dates is not None:
            person_reserve_dates = reserve_scheduled_dates.get(name, set())
            too_close_reserve = [d for d in person_reserve_dates if abs((current_date - d).days) < RESERVE_GAP_DAYS]
            if too_close_reserve:
                if debug_log: rejection_reasons[name] = f"forward reserve_gap: επιφυλακή σε {sorted(too_close_reserve)} (gap<{RESERVE_GAP_DAYS})"
                continue

        eligible.append(p)

    if not eligible:
        if debug_log and rejection_reasons:
            log(f"  DEBUG αποκλεισμοί για {service} {day_name} {current_date}:")
            for n, reason in sorted(rejection_reasons.items()):
                log(f"    {n}: {reason}")
        return None

    eligible.sort(key=lambda person: candidate_priority(service, person, day_name))

    return eligible[start_index % len(eligible)]


def find_reserve_candidate(people, reserve_service, base_service, day_name, current_date, assigned_today, start_index):
    eligible = []

    for p in people:
        if p["energos"] != 1:
            continue

        # πρέπει να μπορεί την επιφυλακή
        if p[reserve_service] != 1:
            continue

        # και να μπορεί και τη βασική υπηρεσία
        if p[base_service] != 1:
            continue

        if p["name"] in assigned_today:
            continue

        # Αν είναι μόνο ΣΚ, δεν μπαίνει καθημερινές
        if p["mono_sk"] == 1 and not is_weekend(day_name):
            continue

        # Αν είναι μόνο 1 τον μήνα + μόνο ΣΚ, κρατάμε προτίμηση για κανονικό ΣΚ
        if is_special_weekend_only_once(p):
            continue

        if p["allowed_days"] and day_name not in p["allowed_days"]:
            continue

        if not posto_allows_service(p, base_service, day_name):
            continue

        if not respects_reserve_gap(p, current_date):
            continue

        eligible.append(p)

    if not eligible:
        return None

    eligible.sort(key=lambda person: reserve_priority(reserve_service, person))

    return eligible[start_index % len(eligible)]


def assign_regular_service(person, current_date, day_name, assigned_today, service=None):
    person["last_service_date"] = current_date
    person["last_service_week"] = get_week_number(current_date)
    person["regular_count"] += 1

    if is_weekend(day_name):
        person["weekend_count"] += 1

    if service:
        count_key = f"{service}_count"
        if count_key in person:
            person[count_key] += 1

    assigned_today.add(person["name"])


def assign_reserve_service(person, current_date, assigned_today):
    person["last_reserve_date"] = current_date
    person["reserve_count"] += 1
    assigned_today.add(person["name"])


def fill_service_slots(
    ws_schedule,
    people,
    row_idx,
    day_name,
    current_date,
    assigned_today,
    service_key,
    slot_names,
    needed_count,
    start_index,
):
    index = start_index

    for slot_idx in range(needed_count):
        if slot_idx >= len(slot_names):
            break

        candidate = find_regular_candidate(
            people=people,
            service=service_key,
            day_name=day_name,
            current_date=current_date,
            assigned_today=assigned_today,
            start_index=index,
        )

        if candidate is None:
            slot_name = slot_names[slot_idx]
            log(f"Generate: κανείς διαθέσιμος για {service_key} slot={slot_name} στις {current_date} — λόγοι:")
            find_regular_candidate(
                people=people, service=service_key, day_name=day_name,
                current_date=current_date, assigned_today=assigned_today,
                start_index=index, debug_log=True
            )
            continue

        slot_name = slot_names[slot_idx]

        ws_schedule.cell(
            row=row_idx,
            column=SERVICE_COLUMNS[slot_name]
        ).value = candidate["name"]

        assign_regular_service(candidate, current_date, day_name, assigned_today, service=service_key)
        index += 1

    return index


def fill_reserve_slot(
    ws_schedule,
    people,
    row_idx,
    day_name,
    current_date,
    assigned_today,
    reserve_slot_name,
    reserve_service,
    base_service,
    needed_count,
    start_index,
):
    index = start_index

    if needed_count <= 0:
        return index

    candidate = find_reserve_candidate(
        people=people,
        reserve_service=reserve_service,
        base_service=base_service,
        day_name=day_name,
        current_date=current_date,
        assigned_today=assigned_today,
        start_index=index,
    )

    if candidate is None:
        return index

    ws_schedule.cell(
        row=row_idx,
        column=SERVICE_COLUMNS[reserve_slot_name]
    ).value = candidate["name"]

    assign_reserve_service(candidate, current_date, assigned_today)
    index += 1

    return index


# -----------------------------
# LIVE STATS WITH EXCEL FORMULAS
# -----------------------------
def write_stats_formulas(ws_stats, ws_personnel, schedule_rows):
    headers = [
        "ΟΝΟΜΑ",
        "ΣΥΝΟΛΟ_ΥΠΗΡΕΣΙΩΝ",
        "ΣΥΝΟΛΟ_ΣΚ",
        "ΜΛΣΝ",
        "ΕΤΑΦ",
        "ΕΣΤ",
        "ΒΑΥΔΜ",
        "ΕΠΙΦ_ΜΛΣΝ",
        "ΕΠΙΦ_ΒΑΥΔΜ",
        "ΣΥΝΟΛΟ_ΕΠΙΦ",
        "ΕΛΕΓΧΟΣ_MAX",
        "ΕΛΕΓΧΟΣ_ΣΚ",
    ]

    # καθαρισμός
    ws_stats.delete_rows(1, ws_stats.max_row)

    # headers με κίτρινο background
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_stats.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = yellow

    start_row = 4
    end_row = schedule_rows[-1]["row_index"]

    day_range = f"Schedule!$A${start_row}:$A${end_row}"
    mlsn1_range = f"Schedule!$C${start_row}:$C${end_row}"
    mlsn2_range = f"Schedule!$D${start_row}:$D${end_row}"
    mlsn3_range = f"Schedule!$E${start_row}:$E${end_row}"
    mlsn4_range = f"Schedule!$F${start_row}:$F${end_row}"
    etaf1_range = f"Schedule!$G${start_row}:$G${end_row}"
    etaf2_range = f"Schedule!$H${start_row}:$H${end_row}"
    est1_range = f"Schedule!$I${start_row}:$I${end_row}"
    est2_range = f"Schedule!$J${start_row}:$J${end_row}"
    vavdm_range = f"Schedule!$K${start_row}:$K${end_row}"
    epif_mlsn_range = f"Schedule!$L${start_row}:$L${end_row}"
    epif_vavdm_range = f"Schedule!$M${start_row}:$M${end_row}"

    last_personnel_row = ws_personnel.max_row
    stats_row = 2

    for personnel_row in range(2, last_personnel_row + 1):
        name = ws_personnel.cell(row=personnel_row, column=1).value

        if name is None or str(name).strip() == "":
            continue

        ws_stats.cell(row=stats_row, column=1).value = str(name).strip()

        # ΜΛΣΝ
        ws_stats.cell(row=stats_row, column=4).value = (
            f'=COUNTIF({mlsn1_range},$A{stats_row})+'
            f'COUNTIF({mlsn2_range},$A{stats_row})+'
            f'COUNTIF({mlsn3_range},$A{stats_row})+'
            f'COUNTIF({mlsn4_range},$A{stats_row})'
        )

        # ΕΤΑΦ
        ws_stats.cell(row=stats_row, column=5).value = (
            f'=COUNTIF({etaf1_range},$A{stats_row})+'
            f'COUNTIF({etaf2_range},$A{stats_row})'
        )

        # ΕΣΤ
        ws_stats.cell(row=stats_row, column=6).value = (
            f'=COUNTIF({est1_range},$A{stats_row})+'
            f'COUNTIF({est2_range},$A{stats_row})'
        )

        # ΒΑΥΔΜ
        ws_stats.cell(row=stats_row, column=7).value = (
            f'=COUNTIF({vavdm_range},$A{stats_row})'
        )

        # ΕΠΙΦ_ΜΛΣΝ
        ws_stats.cell(row=stats_row, column=8).value = (
            f'=COUNTIF({epif_mlsn_range},$A{stats_row})'
        )

        # ΕΠΙΦ_ΒΑΥΔΜ
        ws_stats.cell(row=stats_row, column=9).value = (
            f'=COUNTIF({epif_vavdm_range},$A{stats_row})'
        )

        # ΣΥΝΟΛΟ_ΥΠΗΡΕΣΙΩΝ = μόνο κανονικές
        ws_stats.cell(row=stats_row, column=2).value = f"=SUM(D{stats_row}:G{stats_row})"

        # ΣΥΝΟΛΟ_ΣΚ = μόνο κανονικές σε ΣΚ
        weekend_formula = (
            f'=COUNTIFS({day_range},"Σάββατο",{mlsn1_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Σάββατο",{mlsn2_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Σάββατο",{mlsn3_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Σάββατο",{mlsn4_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Σάββατο",{etaf1_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Σάββατο",{etaf2_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Σάββατο",{est1_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Σάββατο",{est2_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Σάββατο",{vavdm_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Κυριακή",{mlsn1_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Κυριακή",{mlsn2_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Κυριακή",{mlsn3_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Κυριακή",{mlsn4_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Κυριακή",{etaf1_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Κυριακή",{etaf2_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Κυριακή",{est1_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Κυριακή",{est2_range},$A{stats_row})+'
            f'COUNTIFS({day_range},"Κυριακή",{vavdm_range},$A{stats_row})'
        )
        ws_stats.cell(row=stats_row, column=3).value = weekend_formula

        # ΣΥΝΟΛΟ_ΕΠΙΦ
        ws_stats.cell(row=stats_row, column=10).value = f"=SUM(H{stats_row}:I{stats_row})"

        # ΕΛΕΓΧΟΣ_MAX (col M = MAX_ΜΗΝΑ) - VLOOKUP με TRIM για ονόματα με κενά
        ws_stats.cell(row=stats_row, column=11).value = (
            f'=IFERROR(IF(B{stats_row}<=VLOOKUP(TRIM($A{stats_row}),Personnel!$A:$N,13,0),"OK","ΥΠΕΡΒΑΣΗ"),"!")'
        )

        # ΕΛΕΓΧΟΣ_ΣΚ (col N = MAX_ΣΚ)
        ws_stats.cell(row=stats_row, column=12).value = (
            f'=IFERROR(IF(C{stats_row}<=VLOOKUP(TRIM($A{stats_row}),Personnel!$A:$N,14,0),"OK","ΠΟΛΛΑ ΣΚ"),"!")'
        )

        stats_row += 1


def read_adeies(wb):
    """
    Διαβάζει το Adeies φύλλο και επιστρέφει set από (name, date) tuples.
    Στήλες: ΟΝΟΜΑ | ΑΠΟ | ΕΩΣ
    Αν ΕΩΣ είναι κενό, θεωρείται μία μέρα (= ΑΠΟ).
    """
    if "Adeies" not in wb.sheetnames:
        log("Adeies: δεν υπάρχει φύλλο!")
        return set()

    ws_a = wb["Adeies"]
    leave_set = set()

    for row in ws_a.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        try:
            name = str(row[0]).strip()
            date_from = to_python_date(row[1])
            # ΕΩΣ = ημερομηνία επιστροφής (exclusive): αν επιστρέφει 26/3, άδεια μέχρι 25/3
            date_return = to_python_date(row[2]) if row[2] else date_from + timedelta(days=1)
            current = date_from
            while current < date_return:
                leave_set.add((name, current))
                current += timedelta(days=1)
            log(f"Adeies: {name} από {date_from} επιστροφή {date_return}")
        except Exception as e:
            log(f"Adeies ERROR σε γραμμή {row}: {e}")

    log(f"Adeies: σύνολο {len(leave_set)} ημέρες αδειών.")
    return leave_set


def rebuild_state_from_schedule(ws_schedule, schedule_rows, people):
    """Ανακατασκευάζει counts και last dates για κάθε άτομο από το τρέχον schedule."""
    people_by_name = {p["name"]: p for p in people}

    # Αντιστοίχιση col -> service key για per-service counts
    col_to_service = {
        3: "mlsn", 4: "mlsn", 5: "mlsn", 6: "mlsn",
        7: "etaf", 8: "etaf",
        9: "est",  10: "est",
        11: "vavdm",
    }

    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        current_date = row_info["date"]
        day_name = row_info["day_name"]

        for col in REGULAR_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                p = people_by_name.get(str(name).strip())
                if p:
                    p["regular_count"] += 1
                    if is_weekend(day_name):
                        p["weekend_count"] += 1
                    svc = col_to_service.get(col)
                    if svc:
                        p[f"{svc}_count"] += 1
                    p["last_service_date"] = current_date
                    p["last_service_week"] = get_week_number(current_date)

        for col in RESERVE_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                p = people_by_name.get(str(name).strip())
                if p:
                    p["reserve_count"] += 1
                    p["last_reserve_date"] = current_date


def patch_leaves(ws_schedule, schedule_rows, people, needs_by_day, leave_set):
    """
    1. Καθαρίζει τα κελιά των ατόμων που είναι σε άδεια.
    2. Γεμίζει τα άδεια κελιά με αντικαταστάτες.
    """
    # Βήμα 1: Καθάρισμα κελιών λόγω αδειών (παραλείπεται αν δεν υπάρχουν άδειες)
    if leave_set:
        leave_names = {name for (name, _) in leave_set}
        cleared = 0
        for row_info in schedule_rows:
            row_idx = row_info["row_index"]
            current_date = row_info["date"]
            for col in REGULAR_SERVICE_COLS + RESERVE_SERVICE_COLS:
                raw = ws_schedule.cell(row=row_idx, column=col).value
                if raw:
                    name = str(raw).strip()
                    key = (name, current_date)
                    if key in leave_set:
                        log(f"Patch: καθάρισμα {name} στις {current_date} (col {col})")
                        ws_schedule.cell(row=row_idx, column=col).value = None
                        cleared += 1
        log(f"Patch: καθαρίστηκαν {cleared} θέσεις λόγω αδειών.")
    else:
        log("Patch: δεν υπάρχουν άδειες - γέμισμα μόνο χειροκίνητα κενών κελιών.")

    # Βήμα 2: Χτίζουμε scheduled_dates (κανονικές) και reserve_scheduled_dates (επιφυλακές)
    # {name: set of dates} - για έλεγχο gap προς τα μπρος
    scheduled_dates = {}
    reserve_scheduled_dates = {}
    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        current_date = row_info["date"]
        for col in REGULAR_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                n = str(name).strip()
                scheduled_dates.setdefault(n, set()).add(current_date)
        for col in RESERVE_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                n = str(name).strip()
                reserve_scheduled_dates.setdefault(n, set()).add(current_date)

    # Βήμα 3: Pre-compute ΠΛΗΡΕΙΣ μηνιαίοι μετρητές από ολόκληρο το schedule
    # (ώστε max_mina/max_sk να ελέγχονται με βάση τον πλήρη μήνα, όχι μόνο τις
    #  ήδη-επεξεργασμένες μέρες)
    people_by_name = {p["name"]: p for p in people}
    col_to_service = {
        3: "mlsn", 4: "mlsn", 5: "mlsn", 6: "mlsn",
        7: "etaf", 8: "etaf",
        9: "est",  10: "est",
        11: "vavdm",
    }

    for p in people:
        p["regular_count"] = 0
        p["weekend_count"] = 0
        p["reserve_count"] = 0
        p["mlsn_count"] = 0
        p["etaf_count"] = 0
        p["est_count"] = 0
        p["vavdm_count"] = 0

    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        day_name_pre = row_info["day_name"]
        for col in REGULAR_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                p = people_by_name.get(str(name).strip())
                if p:
                    p["regular_count"] += 1
                    if is_weekend(day_name_pre):
                        p["weekend_count"] += 1
                    svc = col_to_service.get(col)
                    if svc:
                        p[f"{svc}_count"] += 1
        for col in RESERVE_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                p = people_by_name.get(str(name).strip())
                if p:
                    p["reserve_count"] += 1

    # Βήμα 4: Single pass - ενημέρωση ΜΟΝΟ last dates + γέμισμα κενών
    # (τα counts ΔΕΝ αλλάζουν από ήδη-ανατεθειμένες θέσεις, μόνο από νέες αναθέσεις)
    # Στο Patch πάντα index=0: παίρνουμε τον πιο δίκαιο υποψήφιο (λιγότερες υπηρεσίες)

    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        day_name = row_info["day_name"]
        current_date = row_info["date"]
        day_needs = needs_by_day.get(day_name, {})

        # Ποιοι είναι ήδη ανατεθειμένοι σήμερα
        assigned_today = set()
        for col in REGULAR_SERVICE_COLS + RESERVE_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                assigned_today.add(str(name).strip())
        # Αποκλείουμε όσους είναι σε άδεια σήμερα
        for (leave_name, leave_date) in leave_set:
            if leave_date == current_date:
                assigned_today.add(leave_name)

        # Ενημέρωση ΜΟΝΟ last dates (counts ήδη pre-computed παραπάνω)
        for col in REGULAR_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                p = people_by_name.get(str(name).strip())
                if p:
                    p["last_service_date"] = current_date
                    p["last_service_week"] = get_week_number(current_date)

        for col in RESERVE_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                p = people_by_name.get(str(name).strip())
                if p:
                    p["last_reserve_date"] = current_date

        # Γέμισμα κενών κανονικών slots
        for service_key, slot_names in REGULAR_SERVICE_SLOTS.items():
            needed = day_needs.get(service_key, 0)
            for slot_idx, slot_name in enumerate(slot_names):
                if slot_idx >= needed:
                    break
                col = SERVICE_COLUMNS[slot_name]
                if ws_schedule.cell(row=row_idx, column=col).value is not None:
                    continue
                candidate = find_regular_candidate(
                    people=people,
                    service=service_key,
                    day_name=day_name,
                    current_date=current_date,
                    assigned_today=assigned_today,
                    start_index=0,
                    scheduled_dates=scheduled_dates,
                    reserve_scheduled_dates=reserve_scheduled_dates,
                )
                if candidate:
                    log(f"Patch: ανάθεση {candidate['name']} -> {slot_name} στις {current_date}")
                    ws_schedule.cell(row=row_idx, column=col).value = candidate["name"]
                    assign_regular_service(candidate, current_date, day_name, assigned_today, service=service_key)
                    scheduled_dates.setdefault(candidate["name"], set()).add(current_date)
                else:
                    log(f"Patch: κανείς διαθέσιμος για {service_key} slot={slot_name} στις {current_date} — λόγοι αποκλεισμού:")
                    find_regular_candidate(
                        people=people, service=service_key, day_name=day_name,
                        current_date=current_date, assigned_today=assigned_today,
                        start_index=0, scheduled_dates=scheduled_dates,
                        reserve_scheduled_dates=reserve_scheduled_dates, debug_log=True
                    )

        # Γέμισμα κενής ΕΠΙΦ_ΜΛΣΝ
        if ws_schedule.cell(row=row_idx, column=12).value is None and day_needs.get("epif_mlsn", 0) > 0:
            candidate = find_reserve_candidate(
                people=people,
                reserve_service="epif_mlsn",
                base_service="mlsn",
                day_name=day_name,
                current_date=current_date,
                assigned_today=assigned_today,
                start_index=0,
            )
            if candidate:
                log(f"Patch: ανάθεση {candidate['name']} -> ΕΠΙΦ_ΜΛΣΝ στις {current_date}")
                ws_schedule.cell(row=row_idx, column=12).value = candidate["name"]
                assign_reserve_service(candidate, current_date, assigned_today)
                reserve_scheduled_dates.setdefault(candidate["name"], set()).add(current_date)

        # Γέμισμα κενής ΕΠΙΦ_ΒΑΥΔΜ
        if ws_schedule.cell(row=row_idx, column=13).value is None and day_needs.get("epif_vavdm", 0) > 0:
            candidate = find_reserve_candidate(
                people=people,
                reserve_service="epif_vavdm",
                base_service="vavdm",
                day_name=day_name,
                current_date=current_date,
                assigned_today=assigned_today,
                start_index=0,
            )
            if candidate:
                log(f"Patch: ανάθεση {candidate['name']} -> ΕΠΙΦ_ΒΑΥΔΜ στις {current_date}")
                ws_schedule.cell(row=row_idx, column=13).value = candidate["name"]
                assign_reserve_service(candidate, current_date, assigned_today)
                reserve_scheduled_dates.setdefault(candidate["name"], set()).add(current_date)

    print("Patch: ολοκληρώθηκε.")


def validate_schedule(wb, ws_schedule, schedule_rows, people, needs_by_day):
    """
    Ελέγχει όλους τους κανόνες μετά το Generate/Patch.
    Γράφει παραβάσεις στο φύλλο 'Violations'.
    """
    if "Violations" in wb.sheetnames:
        ws_v = wb["Violations"]
        ws_v.delete_rows(1, ws_v.max_row)
    else:
        ws_v = wb.create_sheet("Violations")

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red    = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

    headers = ["ΗΜΕΡΟΜΗΝΙΑ", "ΗΜΕΡΑ", "SLOT", "ΟΝΟΜΑ", "ΠΑΡΑΒΑΣΗ"]
    for ci, h in enumerate(headers, 1):
        cell = ws_v.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True)
        cell.fill = yellow
    ws_v.column_dimensions["A"].width = 16   # Ημερομηνία
    ws_v.column_dimensions["B"].width = 14   # Μέρα
    ws_v.column_dimensions["C"].width = 14   # Slot
    ws_v.column_dimensions["D"].width = 32   # Όνομα
    ws_v.column_dimensions["E"].width = 70   # Μήνυμα

    people_by_name = {p["name"]: p for p in people}
    vrow = 2
    violations = 0

    col_to_service = {
        3: "mlsn", 4: "mlsn", 5: "mlsn", 6: "mlsn",
        7: "etaf", 8: "etaf",
        9: "est",  10: "est",
        11: "vavdm",
    }
    col_to_slot = {v: k for k, v in SERVICE_COLUMNS.items()}

    # Συγκεντρώνουμε όλες τις αναθέσεις ανά άτομο
    person_assignments = {}  # name -> [(date, day_name, service, slot)]
    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        current_date = row_info["date"]
        day_name = row_info["day_name"]
        for col in REGULAR_SERVICE_COLS + RESERVE_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                n = str(name).strip()
                svc = col_to_service.get(col, "reserve")
                slot = col_to_slot.get(col, str(col))
                person_assignments.setdefault(n, []).append((current_date, day_name, svc, slot, row_idx, col))

    def add_violation(date, day_name, slot, name, msg):
        nonlocal vrow, violations
        ws_v.cell(row=vrow, column=1, value=date)
        ws_v.cell(row=vrow, column=2, value=day_name)
        ws_v.cell(row=vrow, column=3, value=slot)
        ws_v.cell(row=vrow, column=4, value=name)
        cell = ws_v.cell(row=vrow, column=5, value=msg)
        cell.fill = red
        vrow += 1
        violations += 1
        log(f"VIOLATION: {name} στις {date} ({slot}): {msg}")

    # Έλεγχος κενών slots που έπρεπε να είναι γεμάτα (βάσει Needs)
    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        current_date = row_info["date"]
        day_name = row_info["day_name"]
        day_needs = needs_by_day.get(day_name, {})
        for svc_key, slot_names in REGULAR_SERVICE_SLOTS.items():
            needed = day_needs.get(svc_key, 0)
            for slot_idx, slot_name in enumerate(slot_names):
                if slot_idx >= needed:
                    break
                col = SERVICE_COLUMNS[slot_name]
                if ws_schedule.cell(row=row_idx, column=col).value is None:
                    add_violation(current_date, day_name, slot_name, "—",
                        f"Κενό slot — δεν βρέθηκε υποψήφιος (ή χρειάζεται χειροκίνητη ανάθεση)")
        for epif_key, epif_col, epif_name in [("epif_mlsn", 12, "ΕΠΙΦ_ΜΛΣΝ"), ("epif_vavdm", 13, "ΕΠΙΦ_ΒΑΥΔΜ")]:
            if day_needs.get(epif_key, 0) > 0:
                if ws_schedule.cell(row=row_idx, column=epif_col).value is None:
                    add_violation(current_date, day_name, epif_name, "—",
                        f"Κενό slot — δεν βρέθηκε υποψήφιος (ή χρειάζεται χειροκίνητη ανάθεση)")

    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        current_date = row_info["date"]
        day_name = row_info["day_name"]
        day_needs = needs_by_day.get(day_name, {})
        seen_today = {}

        for col in REGULAR_SERVICE_COLS + RESERVE_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if not name:
                continue
            n = str(name).strip()
            slot = col_to_slot.get(col, str(col))
            svc = col_to_service.get(col, "reserve")
            p = people_by_name.get(n)

            if not p:
                add_violation(current_date, day_name, slot, n, "Δεν βρέθηκε στο Personnel")
                continue

            # Ίδιο άτομο 2 φορές την ίδια μέρα
            if n in seen_today:
                add_violation(current_date, day_name, slot, n, f"Διπλή ανάθεση την ίδια μέρα (και {seen_today[n]})")
            seen_today[n] = slot

            # Ανενεργός στρατιώτης
            if p["energos"] != 1:
                add_violation(current_date, day_name, slot, n, "Ανενεργός στρατιώτης (energos=0) έχει ανατεθεί υπηρεσία")

            # Επιτρεπόμενη υπηρεσία (mlsn/etaf/est/vavdm = 0 → απαγορευμένη)
            if svc in ["mlsn", "etaf", "est", "vavdm"] and not service_allowed_flag(p, svc):
                add_violation(current_date, day_name, slot, n, f"Δεν επιτρέπεται {svc.upper()} (flag=0 στο Personnel)")

            # posto κανόνες
            if not posto_allows_service(p, svc, day_name):
                add_violation(current_date, day_name, slot, n, f"Posto '{p['posto']}' δεν επιτρέπει {svc} ({day_name})")

            # allowed_days
            if p["allowed_days"] and day_name not in p["allowed_days"]:
                add_violation(current_date, day_name, slot, n, f"Επιτρεπτές μέρες: {p['allowed_days']}, βαλμένος σε {day_name}")

            # mono_sk
            if p["mono_sk"] == 1 and not is_weekend(day_name) and col in REGULAR_SERVICE_COLS:
                add_violation(current_date, day_name, slot, n, "MONO_SK αλλά βαλμένος καθημερινή")

        # Υπολογισμός counts ανά άτομο για max checks
    service_counts  = {}  # name -> {svc: count}
    weekend_counts  = {}  # name -> int
    regular_counts  = {}  # name -> int
    reserve_counts  = {}  # name -> int
    service_dates   = {}  # name -> sorted list of (date, type)

    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        current_date = row_info["date"]
        day_name = row_info["day_name"]
        for col in REGULAR_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                n = str(name).strip()
                svc = col_to_service.get(col, "")
                service_counts.setdefault(n, {})
                service_counts[n][svc] = service_counts[n].get(svc, 0) + 1
                regular_counts[n] = regular_counts.get(n, 0) + 1
                if is_weekend(day_name):
                    weekend_counts[n] = weekend_counts.get(n, 0) + 1
                service_dates.setdefault(n, []).append((current_date, "regular"))
        for col in RESERVE_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                n = str(name).strip()
                reserve_counts[n] = reserve_counts.get(n, 0) + 1
                service_dates.setdefault(n, []).append((current_date, "reserve"))

    for n, p in people_by_name.items():
        reg = regular_counts.get(n, 0)
        wk  = weekend_counts.get(n, 0)
        svc_c = service_counts.get(n, {})
        dates = sorted(service_dates.get(n, []), key=lambda x: x[0])

        # max_mina
        if reg > p["max_mina"]:
            add_violation("—", "—", "—", n, f"max_mina υπέρβαση: {reg} υπηρεσίες > επιτρεπτές {p['max_mina']}")

        # max_sk (0 = κανένα ΣΚ επιτρεπτό)
        if wk > p["max_sk"]:
            add_violation("—", "—", "—", n, f"max_sk υπέρβαση: {wk} ΣΚ > επιτρεπτά {p['max_sk']}")

        # mono_1_ton_mina
        if p["mono_1_ton_mina"] == 1 and reg > 1:
            add_violation("—", "—", "—", n, f"MONO_1_TON_MINA: έχει {reg} υπηρεσίες (επιτρεπτή μόνο 1)")

        # per-service max
        for svc in ["mlsn", "etaf", "est", "vavdm"]:
            limit = p.get(f"max_{svc}", 0)
            if limit > 0:
                cnt = svc_c.get(svc, 0)
                if cnt > limit:
                    add_violation("—", "—", svc.upper(), n, f"max_{svc} υπέρβαση: {cnt} > {limit}")

        # min_gap μεταξύ κανονικών υπηρεσιών
        reg_dates = [d for d, t in dates if t == "regular"]
        for i in range(1, len(reg_dates)):
            gap = (reg_dates[i] - reg_dates[i-1]).days
            if gap < p["min_gap_days"]:
                add_violation(reg_dates[i], "—", "—", n,
                    f"min_gap παραβίαση: {gap} μέρες μεταξύ {reg_dates[i-1]} και {reg_dates[i]} (min={p['min_gap_days']})")

        # επιφυλακή gap: δεν επιτρέπεται άλλη υπηρεσία μέσα σε 2 μέρες πριν/μετά την επιφυλακή
        EPIF_GAP = 2
        all_dates = sorted(dates, key=lambda x: x[0])
        reserve_dates = [d for d, t in all_dates if t == "reserve"]
        all_service_dates_only = [d for d, t in all_dates]
        for rd in reserve_dates:
            for d, t in all_dates:
                if d == rd:
                    continue
                gap = abs((d - rd).days)
                if gap <= EPIF_GAP:
                    add_violation(rd, "—", "ΕΠΙΦ", n,
                        f"Επιφυλακή gap παραβίαση: υπηρεσία {'μετά' if d > rd else 'πριν'} {gap} μέρα(ες) ({d})")

    if violations == 0:
        ws_v.cell(row=2, column=1, value="✓ Δεν βρέθηκαν παραβάσεις")
        ws_v.cell(row=2, column=1).fill = PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid")
        log("Validation: όλα OK, 0 παραβάσεις.")
    else:
        log(f"Validation: βρέθηκαν {violations} παραβάσεις!")

    return violations


def save_exclusions(wb, exclusions):
    """
    Αποθηκεύει στο κρυφό φύλλο 'Exclusions' τη λίστα (name, date)
    που έχουν αφαιρεθεί χειροκίνητα. Συσσωρεύεται across πολλαπλά patches.
    """
    if "Exclusions" in wb.sheetnames:
        ws_ex = wb["Exclusions"]
        ws_ex.delete_rows(1, ws_ex.max_row)
    else:
        ws_ex = wb.create_sheet("Exclusions")
        ws_ex.sheet_state = "hidden"
    for i, (name, date) in enumerate(sorted(exclusions, key=lambda x: (x[1], x[0])), start=1):
        ws_ex.cell(row=i, column=1, value=name)
        ws_ex.cell(row=i, column=2, value=date)


def load_exclusions(wb):
    """Διαβάζει το 'Exclusions' φύλλο και επιστρέφει set (name, date)."""
    if "Exclusions" not in wb.sheetnames:
        return set()
    ws_ex = wb["Exclusions"]
    result = set()
    for row in ws_ex.iter_rows(min_row=1, values_only=True):
        if row[0] and row[1]:
            name = str(row[0]).strip()
            date = to_python_date(row[1])
            result.add((name, date))
    return result


def clear_exclusions(wb):
    """Καθαρίζει το Exclusions φύλλο (καλείται από Generate για νέο μήνα)."""
    if "Exclusions" in wb.sheetnames:
        ws_ex = wb["Exclusions"]
        ws_ex.delete_rows(1, ws_ex.max_row)


def save_snapshot(wb, ws_schedule, schedule_rows):
    """
    Σώζει αντίγραφο του τρέχοντος schedule σε κρυφό φύλλο 'Snapshot'.
    Χρησιμοποιείται από το Patch για να εντοπίσει ποιος αφαιρέθηκε χειροκίνητα.
    """
    if "Snapshot" in wb.sheetnames:
        ws_snap = wb["Snapshot"]
        ws_snap.delete_rows(1, ws_snap.max_row)
    else:
        ws_snap = wb.create_sheet("Snapshot")
        ws_snap.sheet_state = "hidden"

    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        snap_row = row_info["day_num"]
        ws_snap.cell(row=snap_row, column=1).value = row_info["date"]
        for col in REGULAR_SERVICE_COLS + RESERVE_SERVICE_COLS:
            val = ws_schedule.cell(row=row_idx, column=col).value
            ws_snap.cell(row=snap_row, column=col).value = val


def load_snapshot_leaves(wb, schedule_rows, ws_schedule):
    """
    Συγκρίνει το τρέχον schedule με το Snapshot.
    Επιστρέφει set (name, date) για κελιά που αδειάστηκαν χειροκίνητα.
    """
    if "Snapshot" not in wb.sheetnames:
        return set()

    ws_snap = wb["Snapshot"]
    manual_leaves = set()

    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        snap_row = row_info["day_num"]
        current_date = row_info["date"]

        for col in REGULAR_SERVICE_COLS + RESERVE_SERVICE_COLS:
            prev_name = ws_snap.cell(row=snap_row, column=col).value
            curr_name = ws_schedule.cell(row=row_idx, column=col).value

            if prev_name and not curr_name:
                name = str(prev_name).strip()
                manual_leaves.add((name, current_date))
                log(f"Snapshot: εντοπίστηκε χειροκίνητη αφαίρεση '{name}' στις {current_date}")

    return manual_leaves


def load_snapshot_manual_additions(wb, schedule_rows, ws_schedule):
    """
    Συγκρίνει το τρέχον schedule με το Snapshot.
    Επιστρέφει set (name, date) για κελιά που ο χρήστης χειροκίνητα ΕΒΑΛΕ
    (ήταν κενά στο Snapshot, τώρα έχουν τιμή — ή άλλαξε το όνομα).
    Χρησιμοποιείται για να αφαιρεθούν από τα Exclusions, ώστε ο χρήστης
    να μπορεί να '赦συγχωρήσει' έναν αποκλεισμό βάζοντας χειροκίνητα το άτομο.
    """
    if "Snapshot" not in wb.sheetnames:
        return set()

    ws_snap = wb["Snapshot"]
    manual_additions = set()

    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        snap_row = row_info["day_num"]
        current_date = row_info["date"]

        for col in REGULAR_SERVICE_COLS + RESERVE_SERVICE_COLS:
            prev_name = ws_snap.cell(row=snap_row, column=col).value
            curr_name = ws_schedule.cell(row=row_idx, column=col).value

            prev = str(prev_name).strip() if prev_name else None
            curr = str(curr_name).strip() if curr_name else None

            if curr and curr != prev:
                # Ήταν κενό ή διαφορετικό → ο χρήστης έβαλε χειροκίνητα
                manual_additions.add((curr, current_date))
                log(f"Snapshot: εντοπίστηκε χειροκίνητη προσθήκη '{curr}' στις {current_date}")

    return manual_additions


def write_carryover(wb, ws_schedule, schedule_rows, people):
    """
    Αποθηκεύει τις τελευταίες N ημέρες του μήνα στο CarryOver φύλλο,
    ώστε το επόμενο generation να ξέρει ποιος υπηρέτησε πρόσφατα.
    N = max(min_gap_days) σε όλους τους στρατιώτες + 1.
    """
    if not people or not schedule_rows:
        return

    max_gap = max(p["min_gap_days"] for p in people)
    days_to_save = max(max_gap, RESERVE_GAP_DAYS) + 1

    if "CarryOver" in wb.sheetnames:
        ws_co = wb["CarryOver"]
        ws_co.delete_rows(1, ws_co.max_row)
    else:
        ws_co = wb.create_sheet("CarryOver")

    ws_co.cell(row=1, column=1).value = "DATE"
    ws_co.cell(row=1, column=2).value = "NAME"
    ws_co.cell(row=1, column=3).value = "TYPE"

    rows_to_save = schedule_rows[-days_to_save:]
    co_row = 2

    for row_info in rows_to_save:
        row_idx = row_info["row_index"]
        current_date = row_info["date"]

        for col in REGULAR_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                ws_co.cell(row=co_row, column=1).value = current_date
                ws_co.cell(row=co_row, column=2).value = name
                ws_co.cell(row=co_row, column=3).value = "regular"
                co_row += 1

        for col in RESERVE_SERVICE_COLS:
            name = ws_schedule.cell(row=row_idx, column=col).value
            if name:
                ws_co.cell(row=co_row, column=1).value = current_date
                ws_co.cell(row=co_row, column=2).value = name
                ws_co.cell(row=co_row, column=3).value = "reserve"
                co_row += 1

    print(f"CarryOver: saved last {days_to_save} days ({co_row - 2} entries).")


def read_carryover(wb, people, schedule_month_year):
    """
    Διαβάζει το CarryOver φύλλο και αρχικοποιεί τα last_service_date /
    last_reserve_date για κάθε στρατιώτη.
    schedule_month_year: (year, month) του μήνα που θα generate-αρουμε,
    ώστε να αγνοούμε CarryOver εγγραφές από τον ίδιο μήνα (διπλό generate).
    """
    if "CarryOver" not in wb.sheetnames:
        print("CarryOver: δεν βρέθηκε φύλλο - ξεκινάμε χωρίς ιστορικό.")
        return

    ws_co = wb["CarryOver"]
    people_by_name = {p["name"]: p for p in people}
    loaded = 0

    for row in ws_co.iter_rows(min_row=2, values_only=True):
        date_val, name, stype = row[0], row[1], row[2]

        if not date_val or not name or not stype:
            continue

        current_date = to_python_date(date_val)

        # Αγνοούμε εγγραφές από τον ίδιο μήνα (προστασία από διπλό generate)
        if (current_date.year, current_date.month) == schedule_month_year:
            continue

        person = people_by_name.get(str(name).strip())
        if not person:
            continue

        if stype == "regular":
            if person["last_service_date"] is None or current_date > person["last_service_date"]:
                person["last_service_date"] = current_date
                person["last_service_week"] = get_week_number(current_date)
                loaded += 1
        elif stype == "reserve":
            if person["last_reserve_date"] is None or current_date > person["last_reserve_date"]:
                person["last_reserve_date"] = current_date
                loaded += 1

    print(f"CarryOver: φορτώθηκαν {loaded} εγγραφές από προηγούμενο μήνα.")


def run_generate(wb, ws_personnel, ws_needs, ws_schedule, ws_stats):
    people = read_personnel(ws_personnel)
    needs_by_day = read_needs(ws_needs)
    schedule_rows = read_schedule_rows(ws_schedule)

    schedule_month_year = (schedule_rows[0]["date"].year, schedule_rows[0]["date"].month)
    read_carryover(wb, people, schedule_month_year)

    print(f"Personnel: {len(people)}, Needs days: {len(needs_by_day)}, Schedule days: {len(schedule_rows)}")

    clear_schedule_services(ws_schedule, schedule_rows)

    index = 0
    reserve_index = 0

    for row_info in schedule_rows:
        row_idx = row_info["row_index"]
        day_name = row_info["day_name"]
        current_date = row_info["date"]
        assigned_today = set()
        day_needs = needs_by_day.get(day_name, {})

        index = fill_service_slots(ws_schedule, people, row_idx, day_name, current_date,
                                   assigned_today, "mlsn", REGULAR_SERVICE_SLOTS["mlsn"],
                                   day_needs.get("mlsn", 0), index)
        index = fill_service_slots(ws_schedule, people, row_idx, day_name, current_date,
                                   assigned_today, "etaf", REGULAR_SERVICE_SLOTS["etaf"],
                                   day_needs.get("etaf", 0), index)
        index = fill_service_slots(ws_schedule, people, row_idx, day_name, current_date,
                                   assigned_today, "vavdm", REGULAR_SERVICE_SLOTS["vavdm"],
                                   day_needs.get("vavdm", 0), index)
        index = fill_service_slots(ws_schedule, people, row_idx, day_name, current_date,
                                   assigned_today, "est", REGULAR_SERVICE_SLOTS["est"],
                                   day_needs.get("est", 0), index)
        reserve_index = fill_reserve_slot(ws_schedule, people, row_idx, day_name, current_date,
                                          assigned_today, "ΕΠΙΦ_ΜΛΣΝ", "epif_mlsn", "mlsn",
                                          day_needs.get("epif_mlsn", 0), reserve_index)
        reserve_index = fill_reserve_slot(ws_schedule, people, row_idx, day_name, current_date,
                                          assigned_today, "ΕΠΙΦ_ΒΑΥΔΜ", "epif_vavdm", "vavdm",
                                          day_needs.get("epif_vavdm", 0), reserve_index)

    write_stats_formulas(ws_stats, ws_personnel, schedule_rows)
    write_carryover(wb, ws_schedule, schedule_rows, people)
    clear_exclusions(wb)
    save_snapshot(wb, ws_schedule, schedule_rows)
    validate_schedule(wb, ws_schedule, schedule_rows, people, needs_by_day)
    log("Generate: ολοκληρώθηκε.")


def run_patch(wb, ws_personnel, ws_needs, ws_schedule, ws_stats):
    people = read_personnel(ws_personnel)
    needs_by_day = read_needs(ws_needs)
    schedule_rows = read_schedule_rows(ws_schedule)

    schedule_month_year = (schedule_rows[0]["date"].year, schedule_rows[0]["date"].month)
    read_carryover(wb, people, schedule_month_year)
    # Δεν κάνουμε rebuild από ολόκληρο τον μήνα - το patch loop χτίζει
    # το state μέρα-μέρα (αλλιώς last_service_date παίρνει μελλοντικές ημερομηνίες)

    leave_set = read_adeies(wb)
    snapshot_leaves = load_snapshot_leaves(wb, schedule_rows, ws_schedule)
    manual_additions = load_snapshot_manual_additions(wb, schedule_rows, ws_schedule)
    # Φόρτωση persistent exclusions (συσσωρευμένες από προηγούμενα patches)
    persistent_exclusions = load_exclusions(wb)
    # Αν ο χρήστης έβαλε χειροκίνητα κάποιον που ήταν αποκλεισμένος → ξεαποκλείεται
    persistent_exclusions -= manual_additions
    # Νέες χειροκίνητες αφαιρέσεις συνδυάζονται με τις ήδη αποθηκευμένες
    all_exclusions = persistent_exclusions | snapshot_leaves
    save_exclusions(wb, all_exclusions)
    combined_leaves = leave_set | all_exclusions
    patch_leaves(ws_schedule, schedule_rows, people, needs_by_day, combined_leaves)
    write_stats_formulas(ws_stats, ws_personnel, schedule_rows)
    write_carryover(wb, ws_schedule, schedule_rows, people)
    save_snapshot(wb, ws_schedule, schedule_rows)
    validate_schedule(wb, ws_schedule, schedule_rows, people, needs_by_day)
    log("run_patch: ολοκληρώθηκε.")


def ensure_adeies_sheet(wb):
    """Δημιουργεί το Adeies φύλλο αν δεν υπάρχει."""
    if "Adeies" not in wb.sheetnames:
        ws_a = wb.create_sheet("Adeies")
        ws_a.cell(row=1, column=1).value = "ΟΝΟΜΑ"
        ws_a.cell(row=1, column=2).value = "ΑΠΟ"
        ws_a.cell(row=1, column=3).value = "ΕΩΣ"
        ws_a.column_dimensions["A"].width = 25
        ws_a.column_dimensions["B"].width = 14
        ws_a.column_dimensions["C"].width = 14
        print("Adeies: δημιουργήθηκε νέο φύλλο.")


BACKUP_DIR = os.path.join(BASE_DIR, "backups")

MONTHS_GR = {
    1: "Ιανουαριος", 2: "Φεβρουαριος", 3: "Μαρτιος",
    4: "Απριλιος",   5: "Μαιος",        6: "Ιουνιος",
    7: "Ιουλιος",   8: "Αυγουστος",    9: "Σεπτεμβριος",
    10: "Οκτωβριος", 11: "Νοεμβριος",  12: "Δεκεμβριος",
}


def backup_excel(mode_label, schedule_month=None, schedule_year=None):
    """Αντιγράφει το Excel σε backups/ πριν κάθε Generate/Patch.
    Όνομα: programma_GENERATE_Μαιος_2026.xlsm — χρησιμοποιεί τον μήνα του Schedule."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    now = datetime.now()
    month = schedule_month if schedule_month else now.month
    year = schedule_year if schedule_year else now.year
    month_gr = MONTHS_GR[month]
    backup_name = f"programma_{mode_label}_{month_gr}_{year}.xlsm"
    backup_path = os.path.join(BACKUP_DIR, backup_name)
    shutil.copy2(EXCEL_FILE, backup_path)
    log(f"Backup: αποθηκεύτηκε → backups/{backup_name}")


def main():
    patch_mode = "--patch" in sys.argv
    mode_label = "PATCH" if patch_mode else "GENERATE"
    print(f"=== Scheduler {mode_label} start ===")

    wb = load_workbook(EXCEL_FILE, keep_vba=True)
    print("Sheets:", wb.sheetnames)

    required_sheets = ["Personnel", "Needs", "Schedule", "Stats"]
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Λείπει το φύλλο: {sheet_name}")

    # Διαβάζουμε τον μήνα του Schedule για σωστό όνομα backup
    try:
        month_start_raw = wb["Schedule"]["B1"].value
        month_start = to_python_date(month_start_raw)
        backup_excel(mode_label, schedule_month=month_start.month, schedule_year=month_start.year)
    except Exception:
        backup_excel(mode_label)

    ensure_adeies_sheet(wb)

    ws_personnel = wb["Personnel"]
    ws_needs = wb["Needs"]
    ws_schedule = wb["Schedule"]
    ws_stats = wb["Stats"]

    # Καθαρισμός κενών από ονόματα Personnel (leading/trailing + διπλά εσωτερικά)
    for r in range(2, ws_personnel.max_row + 1):
        v = ws_personnel.cell(row=r, column=1).value
        if v:
            clean = " ".join(str(v).split())
            if clean != str(v):
                ws_personnel.cell(row=r, column=1).value = clean

    if patch_mode:
        run_patch(wb, ws_personnel, ws_needs, ws_schedule, ws_stats)
    else:
        run_generate(wb, ws_personnel, ws_needs, ws_schedule, ws_stats)

    log("Saving workbook...")
    wb.save(EXCEL_FILE)
    log("Workbook saved OK.")


if __name__ == "__main__":
    # Καθαρισμός log αρχείου στην αρχή κάθε εκτέλεσης
    with open(LOG_FILE, "w", encoding="utf-8") as f:
        f.write(f"=== Scheduler run: {datetime.now()} ===\n")
    try:
        main()
    except Exception as e:
        import traceback
        log(f"CRASH: {e}")
        log(traceback.format_exc())